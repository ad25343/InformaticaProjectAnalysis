"""
STEP 1.5 — Mapping Manifest Agent
Runs immediately after XML parsing, before conversion.

Analyses the graph dict produced by parser_agent and builds a structured
ManifestReport that surfaces every source-to-target connection with a
confidence score, plus all gaps and ambiguous items the reviewer must resolve.

Outputs:
  1. ManifestReport (in-memory, returned to caller)
  2. manifest_<mapping_name>.xlsx  (written to the path supplied by caller)

The xlsx has three sheets:
  • Summary        — one row per mapping: counts and overall status
  • Full Lineage   — every source connection (green=HIGH, amber=MEDIUM, yellow=LOW/UNMAPPED)
  • Review Required — filtered view of LOW + UNMAPPED rows with an editable Override column

When the reviewer fills in the Override column and the xlsx is re-uploaded,
conversion_agent reads it back via load_overrides() and uses those values as
ground truth for any ambiguous or missing connections.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from ..models.schemas import (
    ManifestConfidence, ManifestItem, ManifestItemType,
    ManifestReport, ManifestOverride,
)


# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
_GREEN   = "C6EFCE"   # HIGH confidence
_AMBER   = "FFEB9C"   # MEDIUM confidence
_YELLOW  = "FFC7CE"   # LOW confidence
_RED_TXT = "9C0006"   # LOW text
_BLUE_H  = "1F4E79"   # Header fill (dark blue)
_WHITE   = "FFFFFF"


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def build_manifest(graph: dict) -> ManifestReport:
    """
    Analyse the parser graph dict and return a ManifestReport.
    Does NOT write any files — call write_xlsx() separately.
    """
    items: list[ManifestItem] = []

    mapping_names = [m["name"] for m in graph.get("mappings", [])]
    source_names  = [s["name"] for s in graph.get("sources", [])]
    target_names  = [t["name"] for t in graph.get("targets", [])]

    all_transformations: list[dict] = []
    for m in graph.get("mappings", []):
        all_transformations.extend(m.get("transformations", []))

    all_connectors: list[dict] = []
    for m in graph.get("mappings", []):
        all_connectors.extend(m.get("connectors", []))

    # Pre-build derived sets (mirrors verification_agent logic)
    connected_instances = (
        {c.get("from_instance") for c in all_connectors} |
        {c.get("to_instance")   for c in all_connectors}
    )

    sq_names = {
        t["name"]
        for t in all_transformations
        if t.get("type") == "Source Qualifier"
    }

    sq_connected = {n for n in sq_names if n in connected_instances}

    lookup_source_names: dict[str, str] = {}   # lookup-source-name → LKP transform name
    for t in all_transformations:
        if t.get("type") == "Lookup":
            lkp_src = t.get("table_attribs", {}).get("Lookup table name", "").strip()
            if lkp_src:
                lookup_source_names[lkp_src] = t["name"]

    # ── 1. Source lineage items ───────────────────────────────────────────────
    for src in graph.get("sources", []):
        src_name = src["name"]
        confidence, determination = _score_source(
            src_name, connected_instances, sq_connected, lookup_source_names
        )
        mapping_name = _mapping_for_source(src_name, all_connectors, mapping_names)
        items.append(ManifestItem(
            mapping_name=mapping_name,
            item_type=ManifestItemType.SOURCE_LINEAGE,
            location=src_name,
            description=f"Source '{src_name}' connection to downstream qualifier / lookup",
            tool_determination=determination,
            confidence=confidence,
        ))

    # ── 2. Orphaned ports ─────────────────────────────────────────────────────
    port_from = {(c.get("from_instance"), c.get("from_field")) for c in all_connectors}
    port_to   = {(c.get("to_instance"),   c.get("to_field"))   for c in all_connectors}

    for t in all_transformations:
        ttype = t.get("type", "")
        tname = t.get("name", "")
        mapping_name = _mapping_for_transform(tname, graph)

        for port in t.get("ports", []):
            pname    = port.get("name", "")
            porttype = port.get("porttype", "")

            # Only check OUTPUT ports (or INPUT/OUTPUT passthroughs)
            if "OUTPUT" not in porttype:
                continue

            # Skip well-known internal ports
            if ttype == "Rank" and pname == "RANKINDEX":
                continue
            if ttype in ("Source Qualifier", "Source") and "INPUT" not in porttype:
                pass  # SQ outputs are expected to not feed back up
            if ttype == "Target":
                continue

            key = (tname, pname)
            if key not in port_from:
                items.append(ManifestItem(
                    mapping_name=mapping_name,
                    item_type=ManifestItemType.ORPHANED_PORT,
                    location=f"{tname}.{pname}",
                    description=f"Output port '{pname}' on {ttype} '{tname}' has no downstream connector",
                    tool_determination="No connector found originating from this port",
                    confidence=ManifestConfidence.LOW,
                    notes="Enter IGNORE if intentionally unmapped, or the correct target port name",
                ))

    # ── 3. Lineage gaps — target fields with no incoming connector ────────────
    target_instances = {t["name"] for t in all_transformations if t.get("type") == "Target"}
    for t in all_transformations:
        if t.get("type") != "Target":
            continue
        tname = t["name"]
        mapping_name = _mapping_for_transform(tname, graph)
        for port in t.get("ports", []):
            pname = port.get("name", "")
            key   = (tname, pname)
            if key not in port_to:
                items.append(ManifestItem(
                    mapping_name=mapping_name,
                    item_type=ManifestItemType.LINEAGE_GAP,
                    location=f"{tname}.{pname}",
                    description=f"Target field '{pname}' on '{tname}' has no incoming connector",
                    tool_determination="No connector found feeding this target field",
                    confidence=ManifestConfidence.LOW,
                    notes="Enter source port (e.g. SQ_X.FIELD_Y) or INTENTIONAL_NULL",
                ))

    # ── 4. Expressions (informational — HIGH confidence) ─────────────────────
    for t in all_transformations:
        if t.get("type") != "Expression":
            continue
        tname = t["name"]
        mapping_name = _mapping_for_transform(tname, graph)
        for port in t.get("ports", []):
            expr = port.get("expression", "")
            if expr and expr.strip() not in ("", port.get("name", "")):
                items.append(ManifestItem(
                    mapping_name=mapping_name,
                    item_type=ManifestItemType.EXPRESSION,
                    location=f"{tname}.{port['name']}",
                    description=f"Expression to convert: {expr[:120]}{'…' if len(expr) > 120 else ''}",
                    tool_determination="Will be converted by Claude during conversion step",
                    confidence=ManifestConfidence.HIGH,
                ))

    # ── 5. Lookups (informational — HIGH confidence) ─────────────────────────
    for t in all_transformations:
        if t.get("type") != "Lookup":
            continue
        tname = t["name"]
        mapping_name = _mapping_for_transform(tname, graph)
        lkp_src = t.get("table_attribs", {}).get("Lookup table name", "")
        lkp_cond = t.get("table_attribs", {}).get("Lookup condition", "")
        items.append(ManifestItem(
            mapping_name=mapping_name,
            item_type=ManifestItemType.LOOKUP,
            location=tname,
            description=f"Lookup against '{lkp_src}'" + (f" — condition: {lkp_cond[:80]}" if lkp_cond else ""),
            tool_determination=f"Reference table: {lkp_src}",
            confidence=ManifestConfidence.HIGH,
        ))

    # ── 6. Unresolved parameters ─────────────────────────────────────────────
    for param in graph.get("parameters", []):
        pname = param.get("name", "")
        pval  = param.get("default_value", "")
        items.append(ManifestItem(
            mapping_name=mapping_names[0] if mapping_names else "unknown",
            item_type=ManifestItemType.PARAMETER,
            location=pname,
            description=f"Parameter '{pname}' — default: {pval or '(none)'}",
            tool_determination=f"Default value: {pval or 'not set'}",
            confidence=ManifestConfidence.MEDIUM if pval else ManifestConfidence.LOW,
            notes="Override with environment-specific value if default is wrong",
        ))

    # ── Counts ────────────────────────────────────────────────────────────────
    conf_counts = {c: 0 for c in ManifestConfidence}
    for item in items:
        conf_counts[item.confidence] += 1

    return ManifestReport(
        mapping_names=mapping_names,
        source_count=len(source_names),
        target_count=len(target_names),
        transformation_count=len(all_transformations),
        high_confidence=conf_counts[ManifestConfidence.HIGH],
        medium_confidence=conf_counts[ManifestConfidence.MEDIUM],
        low_confidence=conf_counts[ManifestConfidence.LOW],
        unmapped_count=conf_counts[ManifestConfidence.UNMAPPED],
        review_required=(
            conf_counts[ManifestConfidence.LOW] > 0 or
            conf_counts[ManifestConfidence.UNMAPPED] > 0
        ),
        items=items,
        generated_at=datetime.now(timezone.utc).isoformat(),
    )


def write_xlsx(report: ManifestReport, path: str) -> None:
    """Write the manifest report to an xlsx file at *path*."""
    wb = Workbook()

    _build_summary_sheet(wb, report)
    _build_lineage_sheet(wb, report)
    _build_review_sheet(wb, report)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(path)


def write_xlsx_bytes(report: ManifestReport) -> bytes:
    """Return the xlsx as raw bytes (for in-memory use / API response)."""
    buf = io.BytesIO()
    wb = Workbook()
    _build_summary_sheet(wb, report)
    _build_lineage_sheet(wb, report)
    _build_review_sheet(wb, report)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(buf)
    return buf.getvalue()


def load_overrides(path: str) -> list[ManifestOverride]:
    """
    Read back a reviewer-annotated manifest xlsx and return only the rows
    where the reviewer filled in an Override value.  The conversion agent
    calls this before it begins to resolve ambiguous connections.
    """
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
    except Exception:
        return []

    if "Review Required" not in wb.sheetnames:
        return []

    ws = wb["Review Required"]
    overrides: list[ManifestOverride] = []

    # Find header row — look for "Location" column
    headers: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                headers[cell.value.strip()] = cell.column
        if headers:
            break

    loc_col      = headers.get("Location")
    type_col     = headers.get("Item Type")
    override_col = headers.get("Reviewer Override")
    notes_col    = headers.get("Notes")

    if not (loc_col and override_col):
        return []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            location = row[loc_col - 1]
            item_type_raw = row[type_col - 1] if type_col else None
            override = row[override_col - 1]
            notes    = row[notes_col - 1] if notes_col else None
        except IndexError:
            continue

        if not location or not override:
            continue

        try:
            itype = ManifestItemType(item_type_raw) if item_type_raw else ManifestItemType.SOURCE_LINEAGE
        except ValueError:
            itype = ManifestItemType.SOURCE_LINEAGE

        overrides.append(ManifestOverride(
            location=str(location).strip(),
            item_type=itype,
            reviewer_override=str(override).strip(),
            notes=str(notes).strip() if notes else None,
        ))

    return overrides


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers — confidence scoring
# ─────────────────────────────────────────────────────────────────────────────

def _score_source(
    src_name: str,
    connected_instances: set[str],
    sq_connected: set[str],
    lookup_source_names: dict[str, str],
) -> tuple[ManifestConfidence, str]:
    """Return (confidence, human-readable determination string) for a source."""

    # (a) Direct connector from this source instance
    if src_name in connected_instances:
        return ManifestConfidence.HIGH, f"Direct connector from '{src_name}'"

    # (b) Exact SQ match — SQ_SOURCENAME
    exact_sq = f"SQ_{src_name}"
    if exact_sq in sq_connected:
        return ManifestConfidence.HIGH, f"Exact SQ match → {exact_sq}"

    # (c) Source name contained in some SQ name
    for sq in sq_connected:
        if src_name in sq:
            return ManifestConfidence.HIGH, f"Source name found in SQ name → {sq}"

    # (d) SQ abbreviated name contained in source name  (e.g. APPRAISALS in CORELOGIC_APPRAISALS)
    for sq in sq_connected:
        stem = sq.replace("SQ_", "")
        if stem and stem in src_name:
            return ManifestConfidence.MEDIUM, f"SQ stem '{stem}' found in source name (abbreviated match) → {sq}"

    # (e) Lookup reference table
    if src_name in lookup_source_names:
        lkp = lookup_source_names[src_name]
        return ManifestConfidence.HIGH, f"Lookup reference table → {lkp}"

    # (f) Partial token overlap — weak match
    src_tokens = set(src_name.upper().split("_"))
    for sq in sq_connected:
        sq_tokens = set(sq.replace("SQ_", "").upper().split("_"))
        if src_tokens & sq_tokens:
            return ManifestConfidence.LOW, f"Partial token overlap with SQ '{sq}' — needs review"

    return ManifestConfidence.UNMAPPED, "No matching SQ, direct connector, or Lookup reference found"


def _mapping_for_source(src_name: str, connectors: list[dict], mapping_names: list[str]) -> str:
    for c in connectors:
        if c.get("from_instance") == src_name:
            return mapping_names[0] if mapping_names else "unknown"
    return mapping_names[0] if mapping_names else "unknown"


def _mapping_for_transform(tname: str, graph: dict) -> str:
    for m in graph.get("mappings", []):
        for t in m.get("transformations", []):
            if t["name"] == tname:
                return m["name"]
    mapping_names = [m["name"] for m in graph.get("mappings", [])]
    return mapping_names[0] if mapping_names else "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers — xlsx sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _header_style(cell, text: str) -> None:
    cell.value = cell.value if text is None else text
    cell.font  = Font(name="Arial", bold=True, color=_WHITE, size=10)
    cell.fill  = PatternFill("solid", fgColor=_BLUE_H)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _data_style(cell, fill_hex: Optional[str] = None, bold: bool = False,
                txt_color: str = "000000", wrap: bool = True) -> None:
    cell.font      = Font(name="Arial", size=9, bold=bold, color=txt_color)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_summary_sheet(wb: Workbook, report: ManifestReport) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Informatica Mapping Manifest — Pre-Conversion Review"
    title.font  = Font(name="Arial", bold=True, size=13, color=_WHITE)
    title.fill  = PatternFill("solid", fgColor=_BLUE_H)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = f"Generated: {report.generated_at[:19].replace('T', ' ')} UTC   |   Review Required: {'YES ⚠️' if report.review_required else 'NO ✅'}"
    sub.font  = Font(name="Arial", italic=True, size=9, color="444444")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Stats header row
    headers = ["Mapping", "Sources", "Targets", "Transformations",
               "HIGH ✅", "MEDIUM ⚠", "LOW / UNMAPPED ❌"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        _header_style(cell, h)
    ws.row_dimensions[4].height = 20

    # One data row per mapping (summary level — counts are mapping-level)
    for i, mname in enumerate(report.mapping_names, start=5):
        mapping_items = [x for x in report.items if x.mapping_name == mname]
        mc = {c: sum(1 for x in mapping_items if x.confidence == c) for c in ManifestConfidence}
        low_unmapped = mc[ManifestConfidence.LOW] + mc[ManifestConfidence.UNMAPPED]
        has_issues = low_unmapped > 0

        row_fill = _YELLOW if has_issues else _GREEN
        row_data = [
            mname,
            report.source_count,
            report.target_count,
            report.transformation_count,
            mc[ManifestConfidence.HIGH],
            mc[ManifestConfidence.MEDIUM],
            low_unmapped,
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            _data_style(cell, fill_hex=row_fill, bold=(col == 1))

    _col_widths(ws, [40, 10, 10, 18, 10, 12, 18])


def _build_lineage_sheet(wb: Workbook, report: ManifestReport) -> None:
    ws = wb.create_sheet("Full Lineage")
    ws.sheet_view.showGridLines = False

    headers = ["Mapping", "Item Type", "Location", "Description",
               "Tool Determination", "Confidence", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell, h)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Filter: only SOURCE_LINEAGE and LOOKUP for this sheet (the overview)
    lineage_items = [
        x for x in report.items
        if x.item_type in (
            ManifestItemType.SOURCE_LINEAGE,
            ManifestItemType.LOOKUP,
            ManifestItemType.PARAMETER,
            ManifestItemType.EXPRESSION,
        )
    ]

    for row_idx, item in enumerate(lineage_items, start=2):
        fill = _conf_fill(item.confidence)
        txt  = _RED_TXT if item.confidence in (
            ManifestConfidence.LOW, ManifestConfidence.UNMAPPED
        ) else "000000"

        row_data = [
            item.mapping_name,
            item.item_type.value,
            item.location,
            item.description,
            item.tool_determination,
            item.confidence.value,
            item.notes or "",
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _data_style(cell, fill_hex=fill, txt_color=txt)

    _col_widths(ws, [30, 18, 28, 50, 45, 12, 35])


def _build_review_sheet(wb: Workbook, report: ManifestReport) -> None:
    """
    The action sheet — only LOW/UNMAPPED items from ALL item types.
    Reviewer fills in the 'Reviewer Override' column and saves.
    """
    ws = wb.create_sheet("Review Required")
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:H1")
    banner = ws["A1"]
    banner.value = (
        "⚠️  REVIEWER ACTION REQUIRED — Fill in the 'Reviewer Override' column for each row below, then save and re-upload."
    )
    banner.font  = Font(name="Arial", bold=True, size=10, color="9C0006")
    banner.fill  = PatternFill("solid", fgColor="FFE0E0")
    banner.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = ["Mapping", "Item Type", "Location", "Description",
               "Tool Determination", "Confidence", "Reviewer Override", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        _header_style(cell, h)
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    review_items = [
        x for x in report.items
        if x.confidence in (ManifestConfidence.LOW, ManifestConfidence.UNMAPPED)
    ]

    if not review_items:
        ws.merge_cells("A3:H3")
        cell = ws.cell(row=3, column=1, value="✅  No review required — all items resolved at HIGH or MEDIUM confidence.")
        cell.font      = Font(name="Arial", italic=True, size=9, color="006100")
        cell.fill      = PatternFill("solid", fgColor=_GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        for row_idx, item in enumerate(review_items, start=3):
            fill = _YELLOW if item.confidence == ManifestConfidence.LOW else "FFB3B3"
            row_data = [
                item.mapping_name,
                item.item_type.value,
                item.location,
                item.description,
                item.tool_determination,
                item.confidence.value,
                item.reviewer_override or "",   # ← reviewer fills this in
                item.notes or "",
            ]
            for col, val in enumerate(row_data, start=1):
                bold_override = (col == 7)  # Override column stands out
                cell = ws.cell(row=row_idx, column=col, value=val)
                _data_style(cell, fill_hex=fill, bold=bold_override)

        # Highlight the Override column header extra
        override_header = ws.cell(row=2, column=7)
        override_header.fill = PatternFill("solid", fgColor="C00000")
        override_header.font = Font(name="Arial", bold=True, color=_WHITE, size=10)

    _col_widths(ws, [30, 18, 28, 50, 45, 12, 35, 35])


def _conf_fill(conf: ManifestConfidence) -> str:
    return {
        ManifestConfidence.HIGH:     _GREEN,
        ManifestConfidence.MEDIUM:   _AMBER,
        ManifestConfidence.LOW:      _YELLOW,
        ManifestConfidence.UNMAPPED: "FFB3B3",
    }.get(conf, _WHITE)
