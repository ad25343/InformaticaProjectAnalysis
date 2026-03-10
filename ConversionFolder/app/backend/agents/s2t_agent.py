"""
Source-to-Target Mapping Agent
Runs after Step 1 (Parse). Fully deterministic — no Claude needed.

Traces field lineage through the connector graph to produce:
  - A structured JSON S2T mapping (stored in job state)
  - An Excel workbook (saved to disk, downloadable from UI)

S2T record columns (one row per target field):
  Mapping | Source Table | Source Field | Source Type |
  Transformation Chain | Logic | Logic Type |
  Target Table | Target Field | Target Type |
  Status | Notes
"""
from __future__ import annotations
from pathlib import Path
from typing import Optional
import re

from ..models.schemas import ParseReport


# ── Output directory ─────────────────────────────────────────────────────
APP_DIR      = Path(__file__).parent.parent.parent
S2T_DIR      = APP_DIR / "logs" / "s2t"
S2T_DIR.mkdir(parents=True, exist_ok=True)

# ── Status values ────────────────────────────────────────────────────────
STATUS_DIRECT    = "Direct"        # source field → target with no transformation
STATUS_DERIVED   = "Derived"       # expression / derivation applied
STATUS_FILTERED  = "Filtered"      # passes through a Filter or Router
STATUS_LOOKUP    = "Lookup"        # enriched from a Lookup transformation
STATUS_AGGREGATE = "Aggregated"    # through an Aggregator
STATUS_UNMAPPED_SRC = "Unmapped Source"  # source field has no downstream target
STATUS_UNMAPPED_TGT = "Unmapped Target"  # target field has no upstream source


def build_s2t(
    parse_report: ParseReport,
    graph: dict,
    job_id: str,
) -> dict:
    """
    Build the full S2T mapping for a job.

    Returns a dict:
      {
        "records":        [S2TRecord, ...],       # one per target field
        "unmapped_sources": [...],                # source fields with no target
        "unmapped_targets": [...],                # target fields with no source
        "summary": {...},
        "excel_path": "logs/s2t/<name>.xlsx",     # relative path for download
      }
    """
    records: list[dict] = []
    unmapped_sources: list[dict] = []
    unmapped_targets: list[dict] = []

    # Build fast lookups
    source_lookup  = {s["name"]: s for s in graph.get("sources", [])}
    target_lookup  = {t["name"]: t for t in graph.get("targets", [])}
    source_names   = set(source_lookup.keys())
    target_names   = set(target_lookup.keys())

    for mapping in graph.get("mappings", []):
        mapping_name = mapping["name"]
        connectors   = mapping.get("connectors", [])
        instance_map = mapping.get("instance_map", {})

        # transformation name → transformation dict
        trans_by_name: dict[str, dict] = {
            t["name"]: t for t in mapping.get("transformations", [])
        }
        # Also try instance_map (instance name → actual transform name)
        # so we can look up expressions for instance-named nodes
        def get_trans(inst_name: str) -> Optional[dict]:
            if inst_name in trans_by_name:
                return trans_by_name[inst_name]
            mapped = instance_map.get(inst_name)
            if mapped and mapped in trans_by_name:
                return trans_by_name[mapped]
            return None

        # Build backward index: (to_instance, to_field) → (from_instance, from_field)
        backward: dict[tuple, tuple] = {}
        for conn in connectors:
            key = (conn["to_instance"], conn["to_field"])
            backward[key] = (conn["from_instance"], conn["from_field"])

        # Forward index: (from_instance, from_field) → list of (to_instance, to_field)
        forward: dict[tuple, list] = {}
        for conn in connectors:
            key = (conn["from_instance"], conn["from_field"])
            forward.setdefault(key, []).append((conn["to_instance"], conn["to_field"]))

        # ── Trace each target field backward to its source ────────────────
        target_fields_mapped: set[tuple] = set()

        for tgt in graph.get("targets", []):
            tgt_name = tgt["name"]
            for tgt_field_def in tgt.get("fields", []):
                tgt_field = tgt_field_def["name"]
                tgt_type  = tgt_field_def.get("datatype", "")

                result = _trace_to_source(
                    tgt_name, tgt_field,
                    backward, source_names, trans_by_name, get_trans,
                )

                target_fields_mapped.add((tgt_name, tgt_field))

                if result["source_table"] is None:
                    unmapped_targets.append({
                        "mapping_name": mapping_name,
                        "target_table": tgt_name,
                        "target_field": tgt_field,
                        "target_type":  tgt_type,
                        "note": "No upstream connector found",
                    })
                    continue

                chain_ordered = list(reversed(result["chain"]))
                record = {
                    "mapping_name":        mapping_name,
                    "source_table":        result["source_table"],
                    "source_field":        result["source_field"],
                    "source_type":         _get_source_field_type(
                                              result["source_table"],
                                              result["source_field"],
                                              source_lookup),
                    "transformation_chain": chain_ordered,
                    "transformation_chain_str": " → ".join(chain_ordered) if chain_ordered else "—",
                    "logic":               result["logic"],
                    "logic_type":          result["logic_type"],
                    "target_table":        tgt_name,
                    "target_field":        tgt_field,
                    "target_type":         tgt_type,
                    "status":              result["status"],
                    "notes":               result["notes"],
                }
                records.append(record)

        # ── Find source fields that never reach any target ─────────────────
        # Root nodes = instances that appear as from_instance but never as
        # to_instance. In Informatica these are Source Qualifiers (SQ_*) or
        # occasionally direct source table references.
        to_instances = {conn["to_instance"] for conn in connectors}
        root_instances = {conn["from_instance"] for conn in connectors} - to_instances

        # Collect all (root_instance, field) pairs that appear in from_instance
        # position but are never pulled downstream
        root_fields_used: set[tuple] = {
            (conn["from_instance"], conn["from_field"])
            for conn in connectors
            if conn["from_instance"] in root_instances
        }
        # All (root_instance, field) pairs that actually have a downstream connector
        root_fields_connected: set[tuple] = {
            key for key in forward if key[0] in root_instances
        }
        # Fields present in root but with no downstream = unused source fields
        for (src_inst, src_field) in root_fields_used - root_fields_connected:
            # Try to get datatype from top-level source definition
            # by stripping common SQ prefix conventions (SQ_, SQF_, etc.)
            src_type = ""
            for src in graph.get("sources", []):
                for sf in src.get("fields", []):
                    if sf["name"] == src_field:
                        src_type = sf.get("datatype", "")
                        break
            unmapped_sources.append({
                "mapping_name": mapping_name,
                "source_table": src_inst,
                "source_field": src_field,
                "source_type":  src_type,
                "note": "Field present in source but not connected downstream",
            })

    # ── Build Excel ───────────────────────────────────────────────────────
    mapping_stem = parse_report.mapping_names[0] if parse_report.mapping_names else job_id[:8]
    excel_filename = f"{_safe(mapping_stem)}_s2t_{job_id[:8]}.xlsx"
    excel_path     = S2T_DIR / excel_filename
    _write_excel(records, unmapped_sources, unmapped_targets, excel_path, mapping_stem)

    summary = {
        "total_target_fields":    len(records) + len(unmapped_targets),
        "mapped_fields":          len(records),
        "unmapped_target_fields": len(unmapped_targets),
        "unmapped_source_fields": len(unmapped_sources),
        "direct_mappings":        sum(1 for r in records if r["status"] == STATUS_DIRECT),
        "derived_mappings":       sum(1 for r in records if r["status"] == STATUS_DERIVED),
        "lookup_enriched":        sum(1 for r in records if r["status"] == STATUS_LOOKUP),
        "filtered_fields":        sum(1 for r in records if r["status"] == STATUS_FILTERED),
    }

    return {
        "records":            records,
        "unmapped_sources":   unmapped_sources,
        "unmapped_targets":   unmapped_targets,
        "summary":            summary,
        "excel_filename":     excel_filename,
        "excel_path":         str(excel_path),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Lineage tracer
# ─────────────────────────────────────────────────────────────────────────────

def _trace_to_source(
    start_instance: str,
    start_field: str,
    backward: dict,
    source_names: set,
    trans_by_name: dict,
    get_trans,
    max_depth: int = 20,
) -> dict:
    """
    Walk the backward connector graph from a target field to its ultimate source.
    Returns a dict with: source_table, source_field, chain, logic, logic_type, status, notes.
    """
    chain:  list[str] = []
    logic:  list[str] = []
    notes:  list[str] = []
    status = STATUS_DIRECT

    current_inst  = start_instance
    current_field = start_field
    hops = 0   # track traversal depth to distinguish "no connectors at all" vs "reached root"

    for _ in range(max_depth):
        key = (current_inst, current_field)
        if key not in backward:
            if hops == 0:
                # Target field has NO upstream connector at all → truly unmapped
                return {
                    "source_table": None,
                    "source_field": None,
                    "chain": chain,
                    "logic": "",
                    "logic_type": status,
                    "status": STATUS_UNMAPPED_TGT,
                    "notes": "No upstream connector found",
                }

            # We've hit a dead end after ≥1 hop.  Before declaring this node the "source",
            # check if we're sitting on a derived OUTPUT-only port inside an intermediate
            # transformation (e.g. EXP.ORDER_YEAR_MONTH whose expression references ORDER_DATE).
            # If so, try to follow the expression variable that has a backward connector.
            trans = get_trans(current_inst)
            if trans is not None:
                # Look up the expression for the current port
                expr_text = ""
                for expr in trans.get("expressions", []):
                    if expr["port"] == current_field:
                        expr_text = expr["expression"]
                        break

                if expr_text and expr_text != current_field:
                    # Extract identifiers from the expression and find one that has
                    # an upstream connector at this transformation instance
                    port_names = {p["name"] for p in trans.get("ports", [])}
                    tokens = re.findall(r"\b[A-Za-z_][A-Za-z0-9_]*\b", expr_text)
                    for token in tokens:
                        if token == current_field:
                            continue
                        if token in port_names and (current_inst, token) in backward:
                            # Switch to following this expression input variable
                            notes.append(
                                f"'{current_field}' derived via expression "
                                f"({_truncate(expr_text, 80)}); tracing through '{token}'"
                            )
                            current_field = token
                            break  # restart loop with updated current_field
                    else:
                        # No followable expression variable found — this IS the root
                        return {
                            "source_table": current_inst,
                            "source_field": current_field,
                            "chain": chain,
                            "logic": "; ".join(logic) if logic else "",
                            "logic_type": status,
                            "status": status,
                            "notes": "; ".join(notes) if notes else "",
                        }
                    continue  # re-enter loop with updated current_field

            # No intermediate transformation found (true root / Source Qualifier)
            return {
                "source_table": current_inst,
                "source_field": current_field,
                "chain": chain,
                "logic": "; ".join(logic) if logic else "",
                "logic_type": status,
                "status": status,
                "notes": "; ".join(notes) if notes else "",
            }

        from_inst, from_field = backward[key]
        hops += 1

        if from_inst in source_names:
            # Reached a named source table directly (some Informatica versions
            # connect directly without a Source Qualifier)
            return {
                "source_table": from_inst,
                "source_field": from_field,
                "chain": chain,
                "logic": "; ".join(logic) if logic else "",
                "logic_type": status,
                "status": status,
                "notes": "; ".join(notes) if notes else "",
            }

        # Intermediate transformation node
        chain.append(from_inst)

        trans = get_trans(from_inst)
        if trans:
            ttype = trans.get("type", "")

            # Look for expression on this specific port
            for expr in trans.get("expressions", []):
                if expr["port"] == from_field:
                    expr_text = expr["expression"]
                    if expr_text and expr_text != from_field:
                        logic.append(f"{from_inst}.{from_field}: {_truncate(expr_text, 120)}")
                        status = STATUS_DERIVED

            # Update status based on transformation type
            if ttype == "Lookup Procedure" or "Lookup" in ttype:
                if status == STATUS_DIRECT:
                    status = STATUS_LOOKUP
            elif ttype == "Aggregator":
                if status == STATUS_DIRECT:
                    status = STATUS_AGGREGATE
            elif ttype == "Filter":
                if status == STATUS_DIRECT:
                    status = STATUS_FILTERED
            elif ttype == "Router":
                if status == STATUS_DIRECT:
                    status = STATUS_FILTERED
                notes.append(f"Routes through {from_inst}")
            elif ttype == "Joiner":
                notes.append(f"Joined at {from_inst}")

        current_inst  = from_inst
        current_field = from_field

    # Exceeded max depth
    return {
        "source_table": None,
        "source_field": None,
        "chain": chain,
        "logic": "; ".join(logic),
        "logic_type": status,
        "status": "Trace Too Deep",
        "notes": f"Could not resolve — chain exceeds {max_depth} hops",
    }


def _get_source_field_type(src_table: str, src_field: str, source_lookup: dict) -> str:
    src = source_lookup.get(src_table, {})
    for f in src.get("fields", []):
        if f["name"] == src_field:
            return f.get("datatype", "")
    return ""


def _truncate(s: str, n: int) -> str:
    return s if len(s) <= n else s[:n] + "…"

def _safe(s: str) -> str:
    return re.sub(r"[^\w\-]", "_", s)[:50]


# ─────────────────────────────────────────────────────────────────────────────
# Excel writer
# ─────────────────────────────────────────────────────────────────────────────

def _write_excel(
    records: list[dict],
    unmapped_sources: list[dict],
    unmapped_targets: list[dict],
    path: Path,
    mapping_name: str,
) -> None:
    import openpyxl
    from copy import copy
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── Color palette ─────────────────────────────────────────────────────
    C_HEADER_BG  = "1E293B"   # dark slate
    C_HEADER_FG  = "F8FAFC"   # near white
    C_SRC_BG     = "EFF6FF"   # light blue
    C_TGT_BG     = "F0FDF4"   # light green
    C_DERIVED_BG = "FFFBEB"   # light amber
    C_LOOKUP_BG  = "FAF5FF"   # light purple
    C_FILTER_BG  = "FFF7ED"   # light orange
    C_ERROR_BG   = "FEF2F2"   # light red
    C_DIRECT_BG  = "F0FDF4"   # light green
    C_ALT_BG     = "F8FAFC"   # very light grey (alternating rows)
    C_ACCENT     = "6366F1"   # indigo accent for headers

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_style() -> dict:
        return {
            "font":      Font(bold=True, color=C_HEADER_FG, name="Calibri", size=10),
            "fill":      PatternFill("solid", fgColor=C_HEADER_BG),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border":    border,
        }

    def apply(cell, **kwargs):
        for attr, val in kwargs.items():
            setattr(cell, attr, copy(val))

    def apply_header(cell, value):
        cell.value = value
        hs = header_style()
        apply(cell, **hs)

    STATUS_FILL = {
        STATUS_DIRECT:    PatternFill("solid", fgColor=C_DIRECT_BG),
        STATUS_DERIVED:   PatternFill("solid", fgColor=C_DERIVED_BG),
        STATUS_LOOKUP:    PatternFill("solid", fgColor=C_LOOKUP_BG),
        STATUS_FILTERED:  PatternFill("solid", fgColor=C_FILTER_BG),
        STATUS_AGGREGATE: PatternFill("solid", fgColor=C_LOOKUP_BG),
        STATUS_UNMAPPED_TGT: PatternFill("solid", fgColor=C_ERROR_BG),
        STATUS_UNMAPPED_SRC: PatternFill("solid", fgColor=C_ERROR_BG),
    }

    # ─────────────────────────────────────────────
    # Sheet 1 — Field Mapping (main S2T)
    # ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Field Mapping"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Row 1 — Title
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"Source-to-Target Field Mapping — {mapping_name}"
    title_cell.font      = Font(bold=True, size=13, color=C_ACCENT, name="Calibri")
    title_cell.fill      = PatternFill("solid", fgColor="F1F5F9")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    COLS = [
        ("Mapping",              18),
        ("Source Table",         22),
        ("Source Field",         22),
        ("Source Type",          14),
        ("Transformation Chain", 35),
        ("Logic / Expression",   45),
        ("Logic Type",           14),
        ("Target Table",         22),
        ("Target Field",         22),
        ("Target Type",          14),
        ("Status",               14),
        ("Notes",                35),
    ]

    # Row 2 — Column headers
    for col_idx, (col_name, col_width) in enumerate(COLS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        apply_header(cell, col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[2].height = 22

    # Data rows
    for row_idx, r in enumerate(records, start=3):
        is_alt = (row_idx % 2 == 0)
        default_fill = PatternFill("solid", fgColor=C_ALT_BG if is_alt else "FFFFFF")
        status_fill  = STATUS_FILL.get(r["status"], default_fill)

        values = [
            r["mapping_name"],
            r["source_table"],
            r["source_field"],
            r["source_type"],
            r["transformation_chain_str"],
            r["logic"],
            r["logic_type"],
            r["target_table"],
            r["target_field"],
            r["target_type"],
            r["status"],
            r["notes"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = copy(border)
            cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (5, 6, 12)))
            cell.font      = Font(name="Calibri", size=9)
            # Apply status colour to the Status cell; light fill to src/tgt cols
            if col_idx == 11:  # Status
                cell.fill  = copy(status_fill)
                cell.font  = Font(name="Calibri", size=9, bold=True)
            elif col_idx in (2, 3, 4):  # Source columns
                cell.fill  = PatternFill("solid", fgColor=C_SRC_BG if not is_alt else "EBF4FF")
            elif col_idx in (8, 9, 10):  # Target columns
                cell.fill  = PatternFill("solid", fgColor=C_TGT_BG if not is_alt else "E8FAF0")
            else:
                cell.fill  = copy(default_fill)

        ws.row_dimensions[row_idx].height = 16

    # Auto-filter
    if records:
        ws.auto_filter.ref = f"A2:L{len(records) + 2}"

    # ── Legend block (below the data table) ───────────────────────────────
    legend_row = len(records) + 5
    ws.cell(row=legend_row, column=1).value = "Status Legend"
    ws.cell(row=legend_row, column=1).font  = Font(bold=True, name="Calibri", size=9)
    for i, (status, fill_color) in enumerate([
        (STATUS_DIRECT,    C_DIRECT_BG),
        (STATUS_DERIVED,   C_DERIVED_BG),
        (STATUS_LOOKUP,    C_LOOKUP_BG),
        (STATUS_FILTERED,  C_FILTER_BG),
        (STATUS_AGGREGATE, C_LOOKUP_BG),
        (STATUS_UNMAPPED_TGT, C_ERROR_BG),
    ]):
        cell = ws.cell(row=legend_row + 1 + i, column=1, value=status)
        cell.fill   = PatternFill("solid", fgColor=fill_color)
        cell.border = copy(border)
        cell.font   = Font(name="Calibri", size=9)

    # ─────────────────────────────────────────────
    # Sheet 2 — Unmapped Sources
    # ─────────────────────────────────────────────
    if unmapped_sources:
        ws2 = wb.create_sheet("Unmapped Sources")
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = "A2"

        ws2.merge_cells("A1:E1")
        t2 = ws2["A1"]
        t2.value = "Unmapped Source Fields — these fields exist in the source but are not used in any target mapping"
        t2.font      = Font(bold=True, size=11, color="DC2626", name="Calibri")
        t2.fill      = PatternFill("solid", fgColor="FEF2F2")
        t2.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[1].height = 24

        for col_idx, (col_name, col_width) in enumerate([
            ("Mapping", 22), ("Source Table", 24), ("Source Field", 24),
            ("Source Type", 14), ("Note", 50)
        ], start=1):
            cell = ws2.cell(row=2, column=col_idx)
            apply_header(cell, col_name)
            ws2.column_dimensions[get_column_letter(col_idx)].width = col_width

        for row_idx, r in enumerate(unmapped_sources, start=3):
            vals = [r["mapping_name"], r["source_table"], r["source_field"],
                    r["source_type"], r["note"]]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = copy(border)
                cell.fill      = PatternFill("solid", fgColor="FEF2F2" if row_idx%2==0 else "FFFFFF")
                cell.font      = Font(name="Calibri", size=9)
                cell.alignment = Alignment(vertical="top")

    # ─────────────────────────────────────────────
    # Sheet 3 — Unmapped Targets
    # ─────────────────────────────────────────────
    if unmapped_targets:
        ws3 = wb.create_sheet("Unmapped Targets")
        ws3.sheet_view.showGridLines = False
        ws3.freeze_panes = "A2"

        ws3.merge_cells("A1:E1")
        t3 = ws3["A1"]
        t3.value = "Unmapped Target Fields — these target columns have no mapped source"
        t3.font      = Font(bold=True, size=11, color="DC2626", name="Calibri")
        t3.fill      = PatternFill("solid", fgColor="FEF2F2")
        t3.alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[1].height = 24

        for col_idx, (col_name, col_width) in enumerate([
            ("Mapping", 22), ("Target Table", 24), ("Target Field", 24),
            ("Target Type", 14), ("Note", 50)
        ], start=1):
            cell = ws3.cell(row=2, column=col_idx)
            apply_header(cell, col_name)
            ws3.column_dimensions[get_column_letter(col_idx)].width = col_width

        for row_idx, r in enumerate(unmapped_targets, start=3):
            vals = [r["mapping_name"], r["target_table"], r["target_field"],
                    r["target_type"], r["note"]]
            for col_idx, val in enumerate(vals, start=1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=val)
                cell.border    = copy(border)
                cell.fill      = PatternFill("solid", fgColor="FEF2F2" if row_idx%2==0 else "FFFFFF")
                cell.font      = Font(name="Calibri", size=9)
                cell.alignment = Alignment(vertical="top")

    # ─────────────────────────────────────────────
    # Sheet 4 — Summary
    # ─────────────────────────────────────────────
    ws4 = wb.create_sheet("Summary", 0)   # Insert at position 0 (first tab)
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 20

    ws4.merge_cells("A1:B1")
    s1 = ws4["A1"]
    s1.value     = f"S2T Summary — {mapping_name}"
    s1.font      = Font(bold=True, size=14, color=C_ACCENT, name="Calibri")
    s1.fill      = PatternFill("solid", fgColor="EEF2FF")
    s1.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    summary_rows = [
        ("Total Target Fields",     len(records) + len(unmapped_targets)),
        ("Mapped Fields",           len(records)),
        ("  — Direct",              sum(1 for r in records if r["status"] == STATUS_DIRECT)),
        ("  — Derived",             sum(1 for r in records if r["status"] == STATUS_DERIVED)),
        ("  — Lookup Enriched",     sum(1 for r in records if r["status"] == STATUS_LOOKUP)),
        ("  — Filtered / Routed",   sum(1 for r in records if r["status"] in (STATUS_FILTERED, STATUS_AGGREGATE))),
        ("Unmapped Target Fields",  len(unmapped_targets)),
        ("Unmapped Source Fields",  len(unmapped_sources)),
    ]

    for i, (label, val) in enumerate(summary_rows, start=3):
        lc = ws4.cell(row=i, column=1, value=label)
        vc = ws4.cell(row=i, column=2, value=val)
        lc.font  = Font(name="Calibri", size=10, bold=label.startswith("Total") or label.startswith("Mapped") or label.startswith("Unmapped"))
        vc.font  = Font(name="Calibri", size=10, bold=True)
        row_fill = PatternFill("solid", fgColor="F8FAFC" if i % 2 == 0 else "FFFFFF")
        lc.fill  = row_fill
        vc.fill  = copy(row_fill)
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = copy(border)
        vc.border = copy(border)
        ws4.row_dimensions[i].height = 18

        if label.startswith("Unmapped") and val > 0:
            vc.font = Font(name="Calibri", size=10, bold=True, color="DC2626")

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# Lookup helper — find S2T output path for a job
# ─────────────────────────────────────────────────────────────────────────────

def s2t_excel_path(job_id: str) -> Optional[Path]:
    """Find the S2T Excel file for a given job_id (by short ID suffix)."""
    for f in S2T_DIR.glob(f"*_{job_id[:8]}.xlsx"):
        return f
    return None
